#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para preencher template Excel com dados do calendário JSON
Uso: python calendar_to_excel.py <calendar_json> <template_path> <output_path> <client_name> <month_name> <year>
"""

import sys
import json
import openpyxl
from datetime import datetime
from copy import copy
import os
import calendar
import re
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string, get_column_letter

CALENDAR_FONT_SIZE_PT = 10.0

def get_day_of_week(day, month, year):
    """
    Retorna o dia da semana (0=Segunda, 6=Domingo)
    
    Args:
        day: Dia do mês (1-31)
        month: Mês (1-12)
        year: Ano (ex: 2026)
    
    Returns:
        int: 0=Segunda, 1=Terça, 2=Quarta, 3=Quinta, 4=Sexta, 5=Sábado, 6=Domingo
    """
    date_obj = datetime(year, month, day)
    weekday = date_obj.weekday()  # 0=Monday, 6=Sunday
    return weekday

def format_to_excel(formato):
    """
    Converte formato JSON para formato Excel
    
    Args:
        formato: Formato do post (Static, Reels, Carrossel, Stories, Photos)
    
    Returns:
        str: Formato Excel (Arte | Conteúdo, Reels, etc)
    """
    mapping = {
        'Static': 'Arte | Conteúdo',
        'Reels': 'Reels',
        # No "modelo final.xlsx" o dropdown/CF usa categorias (sem 'Carrossel'/'Stories').
        # Para manter cores/validação funcionando, mapeamos para as opções do template.
        'Carrossel': 'Arte | Conteúdo',
        'Stories': 'Outros',
        'Photos': 'Foto | Institucional',
        'Foto': 'Foto | Institucional',
        'foto': 'Foto | Institucional',
    }
    # Busca case-insensitive: normalizar formato para encontrar no mapeamento
    result = mapping.get(formato)
    if result is None and formato:
        result = mapping.get(formato.capitalize())
    return result or 'Arte | Conteúdo'

def parse_month_name(month_name):
    """
    Converte nome do mês em português para número
    
    Args:
        month_name: Nome do mês (Janeiro, Fevereiro, etc) ou "Janeiro 2026"
    
    Returns:
        int: Número do mês (1-12)
    """
    # Remover ano se presente
    month_only = month_name.split()[0].lower()
    
    months = {
        'janeiro': 1,
        'fevereiro': 2,
        'março': 3,
        'marco': 3,
        'abril': 4,
        'maio': 5,
        'junho': 6,
        'julho': 7,
        'agosto': 8,
        'setembro': 9,
        'outubro': 10,
        'novembro': 11,
        'dezembro': 12
    }
    return months.get(month_only, 1)

def parse_year_from_label(month_label, fallback_year):
    try:
        m = re.search(r"(\d{4})", str(month_label))
        if m:
            return int(m.group(1))
    except Exception:
        pass
    return int(fallback_year)

def _take_first_words(text, max_words=10):
    try:
        s = str(text or '').strip()
        if not s:
            return ''
        words = re.findall(r"\S+", s)
        if len(words) <= int(max_words):
            return ' '.join(words)
        return ' '.join(words[: int(max_words)])
    except Exception:
        return ''

def _extract_day_month_year(date_str):
    """Best-effort parse of many formats.

    Returns: (day, month, year_or_none)
    """
    s = str(date_str or '').strip()
    if not s:
        return (1, None, None)

    # ISO-ish: YYYY-MM-DD (optionally time)
    m = re.search(r"(\d{4})\-(\d{1,2})\-(\d{1,2})", s)
    if m:
        return (int(m.group(3)), int(m.group(2)), int(m.group(1)))

    # YYYY/MM/DD
    m = re.search(r"(\d{4})\/(\d{1,2})\/(\d{1,2})", s)
    if m:
        return (int(m.group(3)), int(m.group(2)), int(m.group(1)))

    # dd/MM[/yyyy] OR MM/DD[/yyyy] (ambiguous). Heuristic:
    # - If second part > 12, treat as MM/DD (month=first, day=second)
    # - Else treat as DD/MM (day=first, month=second)
    m = re.search(r"\b(\d{1,2})\/(\d{1,2})(?:\/(\d{2,4}))?", s)
    if m:
        a = int(m.group(1))
        b = int(m.group(2))
        y = m.group(3)
        year_val = int(y) if (y and len(y) == 4) else None
        if b > 12 and a <= 12:
            return (b, a, year_val)
        return (a, b, year_val)

    # dd-MM[-yyyy] OR MM-DD[-yyyy] (ambiguous). Same heuristic as above.
    m = re.search(r"\b(\d{1,2})\-(\d{1,2})(?:\-(\d{2,4}))?", s)
    if m:
        a = int(m.group(1))
        b = int(m.group(2))
        y = m.group(3)
        year_val = int(y) if (y and len(y) == 4) else None
        if b > 12 and a <= 12:
            return (b, a, year_val)
        return (a, b, year_val)

    # fallback: first number = day
    m = re.search(r"(\d{1,2})", s)
    if m:
        return (int(m.group(1)), None, None)
    return (1, None, None)

def month_num_from_date_str(date_str, default_month):
    try:
        s = str(date_str or '').strip()
        if not s:
            return int(default_month)
        _day, _month, _year = _extract_day_month_year(s)
        if _month is not None and 1 <= int(_month) <= 12:
            print(f"[DEBUG] Parsed date '{date_str}' -> month={_month}, day={_day}")
            return int(_month)
        else:
            print(f"[WARN] Failed to parse month from '{date_str}', using default={default_month}")
    except Exception as e:
        print(f"[ERROR] Exception parsing date '{date_str}': {e}")
    return int(default_month)

def day_from_date_str(date_str):
    try:
        s = str(date_str or '').strip()
        if not s:
            return 1
        _day, _month, _year = _extract_day_month_year(s)
        if 1 <= int(_day) <= 31:
            return int(_day)
        return 1
    except Exception:
        return 1

def month_name_pt(month_num):
    names = {
        1: 'Janeiro',
        2: 'Fevereiro',
        3: 'Março',
        4: 'Abril',
        5: 'Maio',
        6: 'Junho',
        7: 'Julho',
        8: 'Agosto',
        9: 'Setembro',
        10: 'Outubro',
        11: 'Novembro',
        12: 'Dezembro'
    }
    return names.get(int(month_num), 'Janeiro')

def build_post_title(post):
    # Preferir tema; se não houver, usar ideia_visual; fallback: primeira frase do copy
    tema = (post.get('tema') or '').strip()
    if tema:
        return tema

    ideia = (post.get('ideia_visual') or '').strip()
    if ideia:
        return ideia

    copy_text = (post.get('copy_sugestao') or post.get('copy_inicial') or '').strip()
    if not copy_text:
        return ''

    first_line = copy_text.split('\n')[0].strip()
    first_sentence = first_line.split('.') [0].strip() if first_line else ''
    result = first_sentence if first_sentence else first_line
    if len(result) > 90:
        return result[:87].rstrip() + '...'
    return result

def ensure_unique_sheet_name(wb, desired_title):
    title = desired_title
    if title not in wb.sheetnames:
        return title
    suffix = 2
    while True:
        candidate = f"{desired_title} ({suffix})"
        if candidate not in wb.sheetnames:
            return candidate
        suffix += 1

def is_trimestral(periodo):
    try:
        if periodo is None:
            return False
        p = int(str(periodo).strip())
        # Alguns calendários usam 30/60/90 dias; outros podem usar 1/3 meses
        return p >= 90 or p == 3
    except Exception:
        return False

def prepare_month_sheet(wb, base_ws, sheet_title):
    # Remover sheet existente com mesmo nome para evitar '(2)'
    if sheet_title in wb.sheetnames:
        wb.remove(wb[sheet_title])
    ws = wb.copy_worksheet(base_ws)
    ws.title = sheet_title
    return ws

def create_fixed_structure(ws, month_label, client_name):
    title_value = f"{month_label} | {client_name}".strip(" |")
    # Alguns templates têm o título em A1 (ex.: A1:F7), outros em B1 (ex.: B1:G7).
    # Para não deslocar layout, escrevemos no local que já contém o título do template.
    try:
        if ws["A1"].value:
            ws["A1"] = title_value
        elif ws["B1"].value:
            ws["B1"] = title_value
        else:
            ws["A1"] = title_value
    except Exception:
        ws["A1"] = title_value

    # Não sobrescrever a posição do cabeçalho "CHAVE" do template.
    # Alguns templates usam "CHAVE" em outra coluna (ex.: G1) e qualquer escrita fixa aqui pode deslocar o layout.

def fill_single_month(ws, posts, month_num, year_num, client_name, month_label):
    """
    Preenche uma aba com estrutura fixa + blocos semanais de 6 linhas
    Colunas: A=Domingo, B=Segunda, C=Terça, D=Quarta, E=Quinta, F=Sexta, G=Sábado
    """
    # Criar estrutura fixa (linhas 1-8)
    create_fixed_structure(ws, month_label, client_name)

    # Blocos semanais do template: detectamos automaticamente as linhas de cabeçalho das semanas.
    # Isso evita quebrar quando o template muda (offsets/mesclas).

    def _detect_grid_cols():
        # Alguns templates têm grade em A..G, outros em B..H.
        # Detectar olhando a linha de cabeçalho (row 9): se A9 está vazio mas B9 tem texto de dia, usamos B..H.
        try:
            a9 = str(ws["A9"].value or "").strip()
            b9 = str(ws["B9"].value or "").strip()
            if not a9 and b9:
                return ['B', 'C', 'D', 'E', 'F', 'G', 'H']
        except Exception:
            pass
        return ['A', 'B', 'C', 'D', 'E', 'F', 'G']

    col_letters = _detect_grid_cols()

    def _clone_week_block_structure(source_header_row: int, target_header_row: int) -> None:
        block_height = 6
        row_offset = target_header_row - source_header_row
        min_col_idx = column_index_from_string(col_letters[0]) if col_letters else 1
        max_col_idx = column_index_from_string(col_letters[-1]) if col_letters else 7

        for rng in list(ws.merged_cells.ranges):
            if rng.min_row >= target_header_row and rng.max_row <= target_header_row + block_height - 1:
                if rng.min_col >= min_col_idx and rng.max_col <= max_col_idx:
                    ws.unmerge_cells(str(rng))

        for row in range(source_header_row, source_header_row + block_height):
            src_dim = ws.row_dimensions[row]
            dst_dim = ws.row_dimensions[row + row_offset]
            dst_dim.height = src_dim.height
            dst_dim.hidden = src_dim.hidden
            dst_dim.outlineLevel = src_dim.outlineLevel
            dst_dim.collapsed = src_dim.collapsed

            for col_idx in range(min_col_idx, max_col_idx + 1):
                src_cell = ws.cell(row=row, column=col_idx)
                dst_cell = ws.cell(row=row + row_offset, column=col_idx)
                if src_cell.has_style:
                    dst_cell._style = copy(src_cell._style)
                if src_cell.number_format:
                    dst_cell.number_format = src_cell.number_format
                dst_cell.value = None

        for rng in list(ws.merged_cells.ranges):
            if rng.min_row >= source_header_row and rng.max_row <= source_header_row + block_height - 1:
                if rng.min_col >= min_col_idx and rng.max_col <= max_col_idx:
                    target_ref = f"{get_column_letter(rng.min_col)}{rng.min_row + row_offset}:{get_column_letter(rng.max_col)}{rng.max_row + row_offset}"
                    ws.merge_cells(target_ref)

    def _set_font_size(cell_ref: str, size_pt: float = CALENDAR_FONT_SIZE_PT) -> None:
        try:
            font = copy(ws[cell_ref].font)
            font.sz = size_pt
            ws[cell_ref].font = font
        except Exception:
            pass

    def _detect_week_start_rows(max_scan_row: int = 80) -> list:
        # Procura por linhas que tenham cabeçalhos no formato "... (n)".
        # Ex.: "SEGUNDA (1)". Funciona tanto para A..G quanto para B..H.
        rows = []
        try:
            day_header_re = re.compile(r"\((\d{1,2})\)")
            for r in range(1, max_scan_row + 1):
                hits = 0
                for col in col_letters:
                    v = ws[f"{col}{r}"].value
                    if not v:
                        continue
                    s = str(v)
                    if day_header_re.search(s):
                        hits += 1
                # Cabeçalho de semana normalmente tem vários dias na mesma linha
                if hits >= 2:
                    rows.append(r)
        except Exception:
            return []

        # Deduplicar e manter somente o primeiro bloco de 7 semanas (ordem)
        # Precisamos de até 7 semanas para meses que começam no sábado ou domingo
        rows = sorted(set(rows))
        # Alguns templates podem ter hits em linhas que não são calendário; limitar para 6 <= r <= 80
        rows = [r for r in rows if 6 <= r <= 80]

        # Detectar também linhas de semana parcial (quando o mês começa no sábado ou domingo,
        # a primeira linha de semana só tém 1-2 dias, abaixo do limiar de hits >= 2).
        # Buscamos a linha imediatamente anterior à primeira semana completa detectada.
        if rows:
            first_full_row = rows[0]
            partial_re = re.compile(r"\((\d{1,2})\)")
            partial_candidate = None
            # Varrer linhas anteriores à primeira semana completa para um cabeçalho parcial
            for r in range(max(1, first_full_row - 20), first_full_row):
                for col in col_letters:
                    v = ws[f"{col}{r}"].value
                    if v and partial_re.search(str(v)):
                        partial_candidate = r
                        break
                if partial_candidate:
                    break
            if partial_candidate and partial_candidate not in rows:
                rows = [partial_candidate] + rows

        return rows[:7]

    week_start_rows = _detect_week_start_rows()
    if not week_start_rows:
        week_start_rows = [9, 14, 20, 26, 32, 38]

    _, days_in_month = calendar.monthrange(year_num, month_num)

    # Sempre regenerar os cabeçalhos e o mapeamento dia->slot de forma determinística.
    # O template pode vir com textos/posicionamentos diferentes por aba/mês; se tentarmos inferir do texto,
    # qualquer divergência desloca os posts (ex.: dia 2 cai na segunda).
    day_labels = {
        0: 'DOMINGO',
        1: 'SEGUNDA',
        2: 'TER\u00c7A',
        3: 'QUARTA',
        4: 'QUINTA',
        5: 'SEXTA',
        6: 'S\u00c1BADO'
    }

    # Converter de Monday-based para Sunday-based (0=Dom ... 6=Sáb)
    first_weekday_sun0 = (datetime(year_num, month_num, 1).weekday() + 1) % 7

    # Calcular quantas linhas de semana são necessárias para cobrir todos os dias do mês.
    # Um mês que começa no fim de semana pode precisar de até 6 linhas.
    days_in_first_row = (7 - first_weekday_sun0) if first_weekday_sun0 != 0 else 7
    if first_weekday_sun0 == 0:
        weeks_needed = (days_in_month + 6) // 7
    else:
        remaining_after_first = days_in_month - days_in_first_row
        weeks_needed = 1 + (remaining_after_first + 6) // 7 if remaining_after_first > 0 else 1

    # Só sintetizar linhas extras se necessário (caso o template seja mais curto que o mês exige)
    while len(week_start_rows) < weeks_needed:
        if len(week_start_rows) >= 2:
            step = week_start_rows[-1] - week_start_rows[-2]
        else:
            step = 6
        next_row = week_start_rows[-1] + step
        _clone_week_block_structure(week_start_rows[-1], next_row)
        week_start_rows.append(next_row)
        print(f"[INFO] Linha de semana extra sintetizada: row={next_row} (necess?rias={weeks_needed})")

    day_to_slot = {}
    current_day = 1
    for week_index, header_row in enumerate(week_start_rows):
        # Limpar cabeçalhos existentes (valores apenas)
        for col in col_letters:
            try:
                ws[f"{col}{header_row}"].value = None
            except Exception:
                pass

        if current_day > days_in_month:
            continue

        # Semana 0: encaixada quando o dia 1 não começa no domingo.
        if week_index == 0 and first_weekday_sun0 != 0:
            usable_cols = 7 - first_weekday_sun0
            for offset in range(usable_cols):
                col = col_letters[first_weekday_sun0 + offset]
                dow = first_weekday_sun0 + offset
                if current_day <= days_in_month:
                    ws[f"{col}{header_row}"].value = f"{day_labels[dow]} ({current_day})"
                    _set_font_size(f"{col}{header_row}")
                    day_to_slot[current_day] = (header_row, col)
                    current_day += 1
            continue

        # Demais semanas: domingo..sábado em A..G (ou B..H)
        for dow in range(7):
            col = col_letters[dow]
            if current_day <= days_in_month:
                ws[f"{col}{header_row}"].value = f"{day_labels[dow]} ({current_day})"
                _set_font_size(f"{col}{header_row}")
                day_to_slot[current_day] = (header_row, col)
                current_day += 1
            else:
                ws[f"{col}{header_row}"].value = None

    def _top_left_of_merged(cell_ref: str) -> str:
        # Retorna a coordenada top-left do range mesclado que contém cell_ref, ou a própria cell_ref.
        try:
            for rng in ws.merged_cells.ranges:
                if cell_ref in rng:
                    return str(rng.coord).split(":")[0]
        except Exception:
            pass
        return cell_ref

    def _merged_range_for(cell_ref: str):
        # Retorna o range mesclado (objeto) que contém a célula, ou None.
        try:
            for rng in ws.merged_cells.ranges:
                if cell_ref in rng:
                    return rng
        except Exception:
            return None
        return None

    def _content_rows_for_week(header_row: int) -> tuple:
        # Determina dinamicamente as linhas do bloco para:
        # - tipo: header_row+1
        # - resumo: começa em header_row+2 e vai até o fim do merge vertical (se existir)
        # Não retorna a linha separadora (a próxima linha), que deve ser preservada.
        tipo_row = header_row + 1
        resumo_start = header_row + 2

        # Usar a primeira coluna da grade como amostra (o template mescla verticalmente por coluna).
        sample_col = col_letters[0] if col_letters else 'A'
        rng = _merged_range_for(f"{sample_col}{resumo_start}")
        if rng is not None:
            resumo_end = rng.max_row
        else:
            # fallback conservador: 2 linhas de resumo
            resumo_end = resumo_start + 1
        return (tipo_row, resumo_start, resumo_end)

    def _has_conditional_formatting() -> bool:
        try:
            cf = getattr(ws, "conditional_formatting", None)
            rules = getattr(cf, "_cf_rules", None)
            return bool(rules)
        except Exception:
            return False

    def _set_font_size(cell_ref: str, size_pt: float = CALENDAR_FONT_SIZE_PT) -> None:
        try:
            font = copy(ws[cell_ref].font)
            font.sz = size_pt
            ws[cell_ref].font = font
        except Exception:
            pass

    def _legend_style_for_formato(formato_excel: str):
        # Só usado quando o template NÒO tem formatação condicional.
        # Em templates novos (modelo final.xlsx), as cores vêm de conditional formatting.
        try:
            legend_col = 'G' if (col_letters and col_letters[0] == 'A') else 'H'
            legend_map = {
                'Reels': f'{legend_col}2',
                'Arte | Conteúdo': f'{legend_col}3',
                'Foto | Institucional': f'{legend_col}4',
                'Arte | Institucional': f'{legend_col}5',
                'Campanha': f'{legend_col}6',
                'Outros': f'{legend_col}7',
                'Carrossel': f'{legend_col}3',
                'Stories': f'{legend_col}7',
            }
            addr = legend_map.get(formato_excel)
            if not addr:
                return None
            return ws[addr]
        except Exception:
            return None

    def _clear_week_content(header_row: int) -> None:
        # Limpa apenas valores do bloco de conteúdo (tipo + resumo), preservando a linha separadora.
        tipo_row, resumo_start, resumo_end = _content_rows_for_week(header_row)
        for col in col_letters:
            for r in range(tipo_row, resumo_end + 1):
                try:
                    ws[f"{col}{r}"].value = None
                except Exception:
                    pass

    # Limpar células de post (tipo + resumo) mantendo estilos/mesclas
    cleared_weeks = set()
    for d, (header_row, col) in day_to_slot.items():
        if header_row not in cleared_weeks:
            _clear_week_content(header_row)
            cleared_weeks.add(header_row)

    # Preencher posts do JSON nos dias corretos
    total_posts_filled = 0
    for post in posts:
        try:
            day = int(post['day'])
            if day < 1 or day > days_in_month:
                continue

            slot = day_to_slot.get(day)
            if not slot:
                continue

            header_row, col = slot

            formato_excel = format_to_excel(post.get('formato', 'Static'))
            # Posts aqui já chegam normalizados por fill_calendar_template (com a chave 'title').
            # Se chamarmos build_post_title novamente, o resumo pode ficar vazio.
            short_title = (post.get('title') or '').strip() or build_post_title(post)

            # Linha do tipo (formato)
            tipo_cell = _top_left_of_merged(f"{col}{header_row + 1}")
            ws[tipo_cell].value = formato_excel

            # Em templates com conditional formatting (ex.: modelo final.xlsx), NÒO setar fill/font.
            # A cor deve ser aplicada automaticamente pelo Excel baseado no texto.
            if not _has_conditional_formatting():
                try:
                    legend_cell = _legend_style_for_formato(formato_excel)
                    if legend_cell is not None:
                        ws[tipo_cell].fill = copy(legend_cell.fill)
                        ws[tipo_cell].font = copy(legend_cell.font)
                except Exception:
                    pass
            
            _set_font_size(tipo_cell)
            # SEMPRE aplicar alinhamento centralizado (Ctrl+E no Excel) DEPOIS de copiar estilos
            try:
                ws[tipo_cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            except Exception:
                pass

            # Bloco do resumo (3 linhas no template). Normalmente é mesclado verticalmente.
            resumo_cell = _top_left_of_merged(f"{col}{header_row + 2}")
            # Exigência: resumo bem curto (máx 10 palavras)
            base_text = (short_title or post.get('descricao') or post.get('description') or post.get('copy_sugestao') or '').strip()
            ws[resumo_cell].value = _take_first_words(base_text, 10)
            _set_font_size(resumo_cell)
            try:
                ws[resumo_cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            except Exception:
                pass
            total_posts_filled += 1
        except Exception as e:
            print(f"Erro ao processar post: {e}")
            continue

    return total_posts_filled

def fill_calendar_template(calendar_json, template_path, output_path, client_name, month_name, year, periodo=None, selected_months=None):
    """
    Preenche template Excel com dados do calendário JSON
    
    Args:
        calendar_json: Lista de posts do calendário
        template_path: Caminho para modelo final.xlsx
        output_path: Caminho para salvar planilha preenchida
        client_name: Nome do cliente
        month_name: Nome do mês (Janeiro, Fevereiro, etc)
        year: Ano (2026)
    """
    
    print(f"[INFO] Carregando template: {template_path}")
    wb = openpyxl.load_workbook(template_path)
    
    if len(wb.sheetnames) == 0:
        wb.create_sheet("Template")

    base_ws = wb.active

    # Converter mês para número e ano
    start_month_num = parse_month_name(month_name)
    year_num = parse_year_from_label(month_name, year)

    # Normalizar posts e agrupar por (mês, ano)
    posts_by_month_year = {}
    print(f"\n[INFO] Processando {len(calendar_json)} posts do JSON...")
    
    for idx, post in enumerate(calendar_json):
        try:
            # Suporte a ambas as chaves: 'data' (string de data) e 'dia' (número do dia)
            date_str = post.get('data') or ''
            if isinstance(date_str, str) and date_str.strip().lower() in ('undefined', 'null', 'none'):
                date_str = ''
            dia_num = post.get('dia', post.get('day'))
            export_month = post.get('_export_month')
            export_year = post.get('_export_year')
            if not date_str and dia_num is not None:
                # Se s? tem 'dia' (n?mero), converter para string de data usando o m?s expl?cito do post
                try:
                    month_for_post = int(export_month) if export_month is not None else start_month_num
                    year_for_post = int(export_year) if export_year is not None else year_num
                    date_str = f"{int(dia_num)}/{month_for_post}/{year_for_post}"
                except (ValueError, TypeError):
                    date_str = '01/01'
            elif not date_str:
                date_str = '01/01'
            print(f"[DEBUG] Post {idx+1}: data='{date_str}', dia='{dia_num}', formato='{post.get('formato')}'")
            
            # Tentar extrair dia, mês e ano da string de data
            d, m_num_found, y_found = _extract_day_month_year(date_str)
            
            # Month fallback
            m_num = m_num_found if (m_num_found is not None) else start_month_num
            
            # Year fallback logic
            if y_found is not None:
                y_num = y_found
            else:
                # Se o post não tem ano, usamos o rollover baseado no mês inicial
                y_num = year_num
                # S� adiciona 1 ano se houver uma transição clara de final de ano (Nov/Dez -> Jan/Fev/Mar)
                # Ex: Planejamento começa em Dez/2025 e post é de Jan.
                # Se m_num=2 (Fev) e start=3 (Mar), deve permanecer no mesmo ano (2026).
                if m_num < start_month_num:
                    # Heurística: se o mês base é final de ano (>=9) e o post é início de ano (<=4)
                    if start_month_num >= 9 and m_num <= 4:
                        y_num = year_num + 1

            title = build_post_title(post)

            key = (m_num, y_num)
            if key not in posts_by_month_year:
                posts_by_month_year[key] = []

            posts_by_month_year[key].append({
                'day': d,
                'formato': post.get('formato', 'Static'),
                'title': title,
                'descricao': post.get('descricao'),
                'description': post.get('description'),
                'copy_sugestao': post.get('copy_sugestao') or post.get('copy_inicial'),
                'copy_inicial': post.get('copy_inicial'),
                'objetivo': post.get('objetivo'),
                'cta': post.get('cta'),
            })
            print(f"[DEBUG] Post {idx+1} agrupado em {m_num}/{y_num} (dia {d})")
        except Exception as e:
            print(f"[WARN] Erro ao processar post {idx+1}: {e}")
    
    # Lista de chaves detectadas
    detected_keys = sorted(posts_by_month_year.keys(), key=lambda x: (x[1], x[0]))
    print(f"\n[INFO] Meses/Anos detectados: {detected_keys}")

    # Se o usuário selecionou meses, garantimos que eles existam para o ano base
    final_keys = []
    if isinstance(selected_months, list) and len(selected_months) > 0:
        print(f"[INFO] Meses selecionados pelo usuário: {selected_months}")
        for m in selected_months:
            try:
                mm = int(m)
                # Determinar o ano para esse mês selecionado (mesma heurística)
                yy = year_num
                if mm < start_month_num and start_month_num >= 9 and mm <= 4:
                    yy = year_num + 1
                
                key = (mm, yy)
                if key not in final_keys:
                    final_keys.append(key)
                if key not in posts_by_month_year:
                    posts_by_month_year[key] = []
            except Exception:
                continue
    else:
        final_keys = detected_keys

    if not final_keys:
        final_keys = [(start_month_num, year_num)]
        posts_by_month_year[final_keys[0]] = []

    # Ordenar chaves finais cronologicamente a partir do mês inicial
    base_val = year_num * 12 + start_month_num
    final_keys.sort(key=lambda x: (x[1] * 12 + x[0] - base_val) if (x[1] * 12 + x[0] >= base_val) else (9999 + x[1] * 12 + x[0]))

    original_sheets = list(wb.worksheets)
    used_sheets = []

    total_posts_filled = 0
    for idx, (m_num, y_num) in enumerate(final_keys):
        month_label = f"{month_name_pt(m_num)} {y_num}"
        sheet_title = f"{client_name}_{month_name_pt(m_num)}"

        ws = wb.copy_worksheet(base_ws)
        ws.title = ensure_unique_sheet_name(wb, sheet_title)
        used_sheets.append(ws)

        filled = fill_single_month(
            ws,
            posts_by_month_year.get((m_num, y_num), []),
            m_num,
            y_num,
            client_name,
            month_label
        )
        total_posts_filled += filled

    # Remover abas originais do template (mantemos apenas as que geramos neste processo)
    for ws in original_sheets:
        try:
            if ws in wb.worksheets:
                wb.remove(ws)
        except Exception:
            pass

    # Fixar a ordem final das abas conforme used_sheets
    try:
        wb._sheets = list(used_sheets)
    except Exception as e:
        print(f"[WARN] Não foi possível fixar ordem final das abas: {e}")
    
    # Criar diretório de saída se não existir
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"[INFO] Diretório criado: {output_dir}")
    
    # Salvar planilha preenchida
    wb.save(output_path)
    print(f"\n[SUCCESS] Planilha salva: {output_path}")
    print(f"[SUCCESS] Total de posts preenchidos: {total_posts_filled}")
    
    return output_path

def main():
    """Main entry point."""
    use_stdin = len(sys.argv) >= 2 and sys.argv[1] == "--stdin"

    if len(sys.argv) < 7:
        print("Uso: python calendar_to_excel.py <calendar_json|--stdin> <template_path> <output_path> <client_name> <month_name> <year> [periodo] [selected_months_json]")
        sys.exit(1)

    template_path = sys.argv[2]
    output_path = sys.argv[3]
    client_name = sys.argv[4]
    month_name = sys.argv[5]
    year = sys.argv[6]
    periodo = sys.argv[7] if len(sys.argv) >= 8 else None
    selected_months = None
    if len(sys.argv) >= 9:
        try:
            selected_months = json.loads(sys.argv[8])
        except Exception:
            selected_months = None

    if use_stdin:
        calendar_json_str = sys.stdin.buffer.read().decode("utf-8-sig")
    else:
        calendar_source = sys.argv[1]
        if os.path.isfile(calendar_source):
            with open(calendar_source, "r", encoding="utf-8") as calendar_file:
                calendar_json_str = calendar_file.read()
        else:
            calendar_json_str = calendar_source

    try:
        if not calendar_json_str or not calendar_json_str.strip():
            raise ValueError("Nenhum calendario recebido para exportacao")

        calendar_json = json.loads(calendar_json_str)
        print(f"[INFO] Calendario carregado: {len(calendar_json)} posts")

        fill_calendar_template(
            calendar_json,
            template_path,
            output_path,
            client_name,
            month_name,
            year,
            periodo,
            selected_months
        )

        print("\n[SUCCESS] Processo concluido com sucesso!")
        sys.exit(0)

    except Exception as e:
        print(f"[ERROR] Erro fatal: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
if __name__ == "__main__":
    main()

