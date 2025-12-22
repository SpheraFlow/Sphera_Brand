import json
import os
import sys
from datetime import datetime, timezone

import openpyxl
from openpyxl.utils import get_column_letter


def safe_str(value) -> str:
    if value is None:
        return ""
    try:
        return str(value)
    except Exception:
        return repr(value)


def style_to_dict(cell) -> dict:
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    border = cell.border
    protection = cell.protection

    return {
        "number_format": safe_str(cell.number_format),
        "font": {
            "name": font.name,
            "sz": font.sz,
            "b": font.b,
            "i": font.i,
            "u": font.u,
            "color": safe_str(getattr(getattr(font, "color", None), "rgb", None)),
        },
        "fill": {
            "fill_type": fill.fill_type,
            "start_color": safe_str(getattr(getattr(fill, "start_color", None), "rgb", None)),
            "end_color": safe_str(getattr(getattr(fill, "end_color", None), "rgb", None)),
        },
        "alignment": {
            "horizontal": alignment.horizontal,
            "vertical": alignment.vertical,
            "wrap_text": alignment.wrap_text,
            "text_rotation": alignment.text_rotation,
            "shrink_to_fit": alignment.shrink_to_fit,
        },
        "border": {
            "left": getattr(border.left, "style", None),
            "right": getattr(border.right, "style", None),
            "top": getattr(border.top, "style", None),
            "bottom": getattr(border.bottom, "style", None),
        },
        "protection": {
            "locked": protection.locked,
            "hidden": protection.hidden,
        },
    }


def workbook_properties_to_dict(wb) -> dict:
    props = getattr(wb, "properties", None)
    if props is None:
        return {}
    result = {}
    for k in [
        "title",
        "subject",
        "creator",
        "lastModifiedBy",
        "created",
        "modified",
        "description",
        "keywords",
        "category",
        "company",
        "manager",
    ]:
        try:
            v = getattr(props, k, None)
            if v is not None:
                result[k] = safe_str(v)
        except Exception:
            pass
    return result


def merged_top_left_map(ws) -> dict:
    mp = {}
    try:
        for rng in ws.merged_cells.ranges:
            coord = str(rng)
            top_left = coord.split(":")[0]
            for row in range(rng.min_row, rng.max_row + 1):
                for col in range(rng.min_col, rng.max_col + 1):
                    mp[f"{get_column_letter(col)}{row}"] = top_left
    except Exception:
        return {}
    return mp


def page_setup_to_dict(ps) -> dict:
    if ps is None:
        return {}
    result = {}
    for k in [
        "paperSize",
        "orientation",
        "scale",
        "fitToWidth",
        "fitToHeight",
        "usePrinterDefaults",
        "horizontalDpi",
        "verticalDpi",
        "firstPageNumber",
    ]:
        try:
            v = getattr(ps, k, None)
            if v is not None:
                result[k] = v
        except Exception:
            pass
    return result


def page_margins_to_dict(pm) -> dict:
    if pm is None:
        return {}
    result = {}
    for k in [
        "left",
        "right",
        "top",
        "bottom",
        "header",
        "footer",
    ]:
        try:
            v = getattr(pm, k, None)
            if v is not None:
                result[k] = v
        except Exception:
            pass
    return result


def print_options_to_dict(po) -> dict:
    if po is None:
        return {}
    result = {}
    for k in ["gridLines", "headings", "horizontalCentered", "verticalCentered"]:
        try:
            v = getattr(po, k, None)
            if v is not None:
                result[k] = bool(v)
        except Exception:
            pass
    return result


def sheet_view_to_dict(ws) -> dict:
    result = {}
    try:
        views = getattr(ws, "sheet_view", None)
        if views is not None:
            for k in [
                "showGridLines",
                "showRowColHeaders",
                "showZeros",
                "rightToLeft",
                "tabSelected",
                "zoomScale",
                "zoomScaleNormal",
                "zoomScalePageLayoutView",
                "zoomScaleSheetLayoutView",
                "view",
            ]:
                try:
                    v = getattr(views, k, None)
                    if v is not None:
                        result[k] = v
                except Exception:
                    pass
    except Exception:
        pass
    try:
        fp = getattr(ws, "freeze_panes", None)
        if fp is not None:
            result["freeze_panes"] = safe_str(fp)
    except Exception:
        pass
    return result


def sheet_properties_to_dict(ws) -> dict:
    result = {}
    try:
        sp = getattr(ws, "sheet_properties", None)
        if sp is not None:
            try:
                result["tabColor"] = getattr(getattr(sp, "tabColor", None), "rgb", None)
            except Exception:
                pass
            for k in [
                "outlinePr",
                "pageSetUpPr",
                "filterMode",
                "published",
                "codeName",
                "enableFormatConditionsCalculation",
            ]:
                try:
                    v = getattr(sp, k, None)
                    if v is not None:
                        result[k] = safe_str(v)
                except Exception:
                    pass
    except Exception:
        pass
    return result


def sheet_protection_to_dict(ws) -> dict:
    prot = getattr(ws, "protection", None)
    if prot is None:
        return {}
    result = {}
    for k in [
        "sheet",
        "objects",
        "scenarios",
        "formatCells",
        "formatColumns",
        "formatRows",
        "insertColumns",
        "insertRows",
        "insertHyperlinks",
        "deleteColumns",
        "deleteRows",
        "selectLockedCells",
        "selectUnlockedCells",
        "sort",
        "autoFilter",
        "pivotTables",
    ]:
        try:
            v = getattr(prot, k, None)
            if v is not None:
                result[k] = bool(v)
        except Exception:
            pass
    try:
        result["password"] = safe_str(getattr(prot, "password", None))
    except Exception:
        pass
    return result


def auto_filter_to_dict(ws) -> dict:
    af = getattr(ws, "auto_filter", None)
    if af is None:
        return {}
    result = {}
    try:
        result["ref"] = safe_str(getattr(af, "ref", None))
    except Exception:
        pass
    try:
        flt_cols = []
        for fc in getattr(af, "filterColumn", []) or []:
            try:
                flt_cols.append(
                    {
                        "colId": getattr(fc, "colId", None),
                        "hiddenButton": getattr(fc, "hiddenButton", None),
                        "showButton": getattr(fc, "showButton", None),
                    }
                )
            except Exception:
                continue
        if flt_cols:
            result["filterColumn"] = flt_cols
    except Exception:
        pass
    return result


def tables_to_list(ws) -> list:
    tables = []
    try:
        tbls = getattr(ws, "tables", None)
        if tbls is not None:
            for t in list(tbls.values()):
                try:
                    tables.append(
                        {
                            "name": getattr(t, "name", None),
                            "displayName": getattr(t, "displayName", None),
                            "ref": getattr(t, "ref", None),
                            "tableStyleInfo": safe_str(getattr(t, "tableStyleInfo", None)),
                            "totalsRowShown": getattr(t, "totalsRowShown", None),
                        }
                    )
                except Exception:
                    continue
    except Exception:
        pass
    return tables


def conditional_formatting_to_list(ws) -> list:
    out = []
    try:
        cf = getattr(ws, "conditional_formatting", None)
        if cf is None:
            return out
        rules_map = getattr(cf, "_cf_rules", None)
        if not rules_map:
            return out
        for sqref, rules in rules_map.items():
            item = {"sqref": safe_str(sqref), "rules": []}
            for r in rules or []:
                try:
                    dxf = getattr(r, "dxf", None)
                    dxf_dict = None
                    if dxf is not None:
                        try:
                            dxf_dict = {
                                "font": {
                                    "color": safe_str(getattr(getattr(getattr(dxf, "font", None), "color", None), "rgb", None)),
                                    "b": getattr(getattr(dxf, "font", None), "b", None),
                                    "i": getattr(getattr(dxf, "font", None), "i", None),
                                },
                                "fill": {
                                    "patternType": getattr(getattr(dxf, "fill", None), "patternType", None),
                                    "start_color": safe_str(getattr(getattr(getattr(dxf, "fill", None), "fgColor", None), "rgb", None)),
                                    "end_color": safe_str(getattr(getattr(getattr(dxf, "fill", None), "bgColor", None), "rgb", None)),
                                },
                                "numFmt": safe_str(getattr(dxf, "numFmt", None)),
                            }
                        except Exception:
                            dxf_dict = safe_str(dxf)

                    item["rules"].append(
                        {
                            "type": getattr(r, "type", None),
                            "priority": getattr(r, "priority", None),
                            "operator": getattr(r, "operator", None),
                            "stopIfTrue": getattr(r, "stopIfTrue", None),
                            "text": getattr(r, "text", None),
                            "rank": getattr(r, "rank", None),
                            "stdDev": getattr(r, "stdDev", None),
                            "timePeriod": getattr(r, "timePeriod", None),
                            "formula": [safe_str(f) for f in (getattr(r, "formula", None) or [])],
                            "dxf": dxf_dict,
                        }
                    )
                except Exception:
                    continue
            out.append(item)
    except Exception:
        pass
    return out


def dump_workbook_overview(wb, out_dir: str) -> None:
    overview = {
        "generated_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "sheetnames": list(wb.sheetnames),
        "defined_names": [],
        "properties": workbook_properties_to_dict(wb),
    }

    try:
        for dn in wb.defined_names.definedName:
            overview["defined_names"].append(
                {
                    "name": dn.name,
                    "attr_text": dn.attr_text,
                    "local_sheet_id": dn.localSheetId,
                    "hidden": dn.hidden,
                }
            )
    except Exception:
        pass

    with open(os.path.join(out_dir, "workbook_overview.json"), "w", encoding="utf-8") as f:
        json.dump(overview, f, ensure_ascii=False, indent=2)


def dump_sheet(ws, out_dir: str, max_rows: int, max_cols: int) -> None:
    sheet_dir = os.path.join(out_dir, f"sheet_{ws.title}")
    os.makedirs(sheet_dir, exist_ok=True)

    merge_map = merged_top_left_map(ws)

    summary = {
        "title": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_cells": sorted([str(rng) for rng in ws.merged_cells.ranges]),
        "row_dimensions": {},
        "column_dimensions": {},
        "data_validations": [],
        "conditional_formatting_rules": 0,
        "conditional_formatting": [],
        "page_setup": page_setup_to_dict(getattr(ws, "page_setup", None)),
        "page_margins": page_margins_to_dict(getattr(ws, "page_margins", None)),
        "print_options": print_options_to_dict(getattr(ws, "print_options", None)),
        "sheet_view": sheet_view_to_dict(ws),
        "sheet_properties": sheet_properties_to_dict(ws),
        "sheet_protection": sheet_protection_to_dict(ws),
        "auto_filter": auto_filter_to_dict(ws),
        "tables": tables_to_list(ws),
    }

    # Row/column dimensions (somente as definidas explicitamente)
    for r, dim in ws.row_dimensions.items():
        if r is None:
            continue
        summary["row_dimensions"][str(r)] = {
            "height": dim.height,
            "hidden": dim.hidden,
            "outlineLevel": dim.outlineLevel,
        }

    for col, dim in ws.column_dimensions.items():
        summary["column_dimensions"][str(col)] = {
            "width": dim.width,
            "hidden": dim.hidden,
            "outlineLevel": dim.outlineLevel,
        }

    # Data validations
    try:
        dvs = getattr(ws, "data_validations", None)
        if dvs is not None:
            for dv in dvs.dataValidation:
                summary["data_validations"].append(
                    {
                        "type": dv.type,
                        "formula1": dv.formula1,
                        "formula2": dv.formula2,
                        "allow_blank": dv.allowBlank,
                        "show_drop_down": getattr(dv, "showDropDown", None),
                        "sqref": safe_str(dv.sqref),
                    }
                )
    except Exception:
        pass

    # Conditional formatting (count only)
    try:
        cf = getattr(ws, "conditional_formatting", None)
        if cf is not None:
            # openpyxl guarda regras internamente; contar com melhor esforço
            count = 0
            for _ in cf._cf_rules.values():
                count += 1
            summary["conditional_formatting_rules"] = count
    except Exception:
        pass

    try:
        summary["conditional_formatting"] = conditional_formatting_to_list(ws)
    except Exception:
        summary["conditional_formatting"] = []

    with open(os.path.join(sheet_dir, "summary.json"), "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    # Dump de células (primeiros max_rows/max_cols)
    cells_dump = []
    for r in range(1, max_rows + 1):
        for c in range(1, max_cols + 1):
            cell = ws.cell(row=r, column=c)
            addr = f"{get_column_letter(c)}{r}"
            try:
                hyperlink = getattr(getattr(cell, "hyperlink", None), "target", None)
            except Exception:
                hyperlink = None
            try:
                comment_text = safe_str(getattr(getattr(cell, "comment", None), "text", None))
            except Exception:
                comment_text = ""
            try:
                is_formula = bool(getattr(cell, "data_type", None) == "f") or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                )
            except Exception:
                is_formula = False
            cells_dump.append(
                {
                    "address": addr,
                    "row": r,
                    "col": c,
                    "value": cell.value,
                    "data_type": cell.data_type,
                    "is_formula": is_formula,
                    "hyperlink": hyperlink,
                    "comment": comment_text,
                    "merged_top_left": merge_map.get(addr),
                    "style": style_to_dict(cell),
                }
            )

    with open(os.path.join(sheet_dir, "cells_dump.json"), "w", encoding="utf-8") as f:
        json.dump(cells_dump, f, ensure_ascii=False, indent=2, default=safe_str)

    # Preview compacto no console (primeiras linhas, valores apenas)
    print(f"\nAba: {ws.title}")
    print(f"Dimensões: max_row={ws.max_row}, max_col={ws.max_column}")
    print(f"Merged cells: {len(summary['merged_cells'])}")
    print(f"Data validations: {len(summary['data_validations'])}")
    print(f"Conditional formatting rules: {summary['conditional_formatting_rules']}")
    print("Preview (valores):")
    headers = [get_column_letter(c) for c in range(1, max_cols + 1)]
    print("   | " + " | ".join(headers))
    for r in range(1, min(max_rows, 12) + 1):
        row_vals = [safe_str(ws.cell(r, c).value) for c in range(1, max_cols + 1)]
        print(f"{str(r).rjust(2)} | " + " | ".join(row_vals))


def inspect_xlsx(xlsx_path: str, out_dir: str, max_rows: int = 60, max_cols: int = 12) -> None:
    print("=")
    print(f"Arquivo: {xlsx_path}")

    if not os.path.exists(xlsx_path):
        print("[ERRO] Arquivo não encontrado.")
        return

    wb = openpyxl.load_workbook(xlsx_path)
    os.makedirs(out_dir, exist_ok=True)

    dump_workbook_overview(wb, out_dir)

    print("Abas:")
    for name in wb.sheetnames:
        print(f"- {name}")

    for name in wb.sheetnames:
        ws = wb[name]
        dump_sheet(ws, out_dir, max_rows=max_rows, max_cols=max_cols)

    print(f"\n[OK] Dumps gerados em: {out_dir}")


def main() -> None:
    repo_root = os.path.dirname(os.path.abspath(__file__))

    # Uso:
    #   python debug_planilha.py <xlsx_path> [out_dir] [max_rows] [max_cols]
    # Se nenhum argumento for passado, inspeciona o template padrão.

    if len(sys.argv) >= 2 and sys.argv[1].strip():
        xlsx_path = sys.argv[1]
        out_dir = sys.argv[2] if len(sys.argv) >= 3 and sys.argv[2].strip() else os.path.join(repo_root, "debug_output")
        try:
            max_rows = int(sys.argv[3]) if len(sys.argv) >= 4 else 60
        except Exception:
            max_rows = 60
        try:
            max_cols = int(sys.argv[4]) if len(sys.argv) >= 5 else 12
        except Exception:
            max_cols = 12

        inspect_xlsx(xlsx_path, out_dir=out_dir, max_rows=max_rows, max_cols=max_cols)
        return

    out_dir = os.path.join(repo_root, "debug_output")
    tri_path = os.path.join(repo_root, "calendario", "CoreSport_Tri_2026.xlsx")
    inspect_xlsx(tri_path, out_dir=os.path.join(out_dir, "CoreSport_Tri_2026"), max_rows=60, max_cols=12)


if __name__ == "__main__":
    main()
